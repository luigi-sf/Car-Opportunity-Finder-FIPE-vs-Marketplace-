import pandas as pd
from excel.utils.open_ia import extrair_com_ia
from excel.utils.fipe import buscar_fipe_cache
from excel.utils.email_sender import enviar_email
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook


def extrair_marca_modelo(titulo):
    try:
        partes = titulo.split()

        marca = partes[0]
        modelo = partes[1]

        if len(marca) < 3 or len(modelo) < 2:
            return extrair_com_ia(titulo)

        return marca, modelo

    except:
        return extrair_com_ia(titulo)


def analisar():
    df = pd.read_excel("carros.xlsx")

    oportunidades = []

    for _, row in df.iterrows():
        titulo = row["Titulo"]
        preco = row["Preco"]
        ano = row["Ano"]
        km = row["KM"]
        link = row["Link"]

        if pd.isna(preco) or pd.isna(ano) or pd.isna(km):
            continue

        marca, modelo = extrair_marca_modelo(titulo)

        if not isinstance(marca, str) or not isinstance(modelo, str):
            continue

        fipe = buscar_fipe_cache(marca, modelo, int(ano))

        if not fipe:
            continue

        desconto = (fipe - preco) / fipe

        if preco < fipe:
            if desconto >= 0.20:
                nivel = "🔥 imperdível"
            elif desconto >= 0.10:
                nivel = "💰 boa"
            else:
                nivel = "⚠️ leve"

            oportunidades.append({
                "Titulo": titulo,
                "Preco": preco,
                "FIPE": fipe,
                "Desconto (%)": desconto,
                "Ano": ano,
                "KM": km,
                "nivel": nivel,
                "Link": link
            })

    if not oportunidades:
        print("Nenhuma oportunidade encontrada 😢")
        return None, 0

    df_final = pd.DataFrame(oportunidades)

    prioridade = {
        "🔥 imperdível": 0,
        "💰 boa": 1,
        "⚠️ leve": 2
    }

    df_final["prioridade"] = df_final["nivel"].map(prioridade)

    df_final = df_final.sort_values(
        by=["prioridade", "Desconto (%)"],
        ascending=[True, False]
    )

    df_final.drop(columns=["prioridade"], inplace=True)

    arquivo = "oportunidades.xlsx"
    df_final.to_excel(arquivo, index=False)

    # FORMATAÇÃO EXCEL
    wb = load_workbook(arquivo)
    ws = wb.active

    # Formatação dos valores
    for row in ws.iter_rows(min_row=2):
        row[1].number_format = 'R$ #,##0.00'   # Preço
        row[2].number_format = 'R$ #,##0.00'   # FIPE
        row[3].number_format = '0.00%'         # Desconto
        row[5].number_format = '#,##0 "km"'    # KM

    # Auto ajuste de largura
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)

        max_length = 0
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        largura = max(max_length + 5, 15)

        # ajustes específicos
        if col_letter in ["B", "C"]:  # Preço e FIPE
            largura = max(largura, 18)

        if col_letter == "F":  # KM
            largura = max(largura, 12)

        ws.column_dimensions[col_letter].width = largura

    wb.save(arquivo)

    total = len(oportunidades)

    print(f"{total} oportunidades encontradas 🔥")

    return arquivo, total


if __name__ == "__main__":
    arquivo, total = analisar()

    if arquivo:
        enviar_email(arquivo, total)